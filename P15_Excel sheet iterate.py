import openpyxl

wb = openpyxl.load_workbook('data sheet.xlsx')

sheets = wb.active

max_row = sheets.max_row

for i in range(1, max_row+1):
    cell = sheets.cell(row=i, column=1)
    print(cell.value)
for i in range(1, max_row+1):
    cell = sheets.cell(row=i, column=2)
    print(cell.value)
.............................................
import pyautogui as pg
import webbrowser as web
import time
import openpyxl

wb = openpyxl.load_workbook('data sheet.xlsx')

sheets = wb.active

max_row = sheets.max_row

for i in range(1, max_row+1):
    cell1 = sheets.cell(row=i, column=1)
   # print(cell1.value)
for i in range(1, max_row+1):
    cell2 = sheets.cell(row=i, column=1)
   # print(cell2.value)

ccol = sheets.ncols
crow = sheets.nrows

#data_dict = data.to_dic("list")
leads = cell1.value
message = cell2.value

#combo = zip(Lead, message)

first = True

time.sleep(4)
for i in range(1, ccol):
    web.open('https://web.whatsapp.com/send?phone='+cell1.value+"&text="+cell2.value)
    if first :
        time.sleep(6)
        first = False
    width.height = pg.size()
    pg.click(width/2,height/2)
    time.sleep(8)
    pg.press('Enter')
    time.sleep(8)
    pg.hotkey('ctrl', 'w')
